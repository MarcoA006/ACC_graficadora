"""
Dashboard de Ventas de Chips ATT
Requiere: pip install customtkinter pandas openpyxl matplotlib pillow
"""

import customtkinter as ctk
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import threading
import gc

# ── Fix: CTkScrollableFrame bug con Python 3.14 ───────────────────────────────
try:
    from customtkinter.windows.widgets.ctk_scrollable_frame import CTkScrollableFrame as _CTKSF
    _orig_check = _CTKSF.check_if_master_is_canvas
    def _safe_check(self, widget):
        try:
            return _orig_check(self, widget)
        except AttributeError:
            return False
    _CTKSF.check_if_master_is_canvas = _safe_check
except Exception:
    pass

# ── Tema ──────────────────────────────────────────────────────────────────────
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

ACCENT   = "#1F6AA5"
ACCENT2  = "#E87B2E"
BG_DARK  = "#1a1a2e"
BG_MID   = "#16213e"
BG_CARD  = "#0f3460"
TEXT_CLR = "#e0e0e0"
MONTH_COLORS = ["#4e9af1","#f18f4e","#4ef1a0","#f14e4e","#c44ef1","#f1d84e"]

MESES_ES = {1:"ene", 2:"feb", 3:"mar", 4:"abr", 5:"may", 6:"jun",
            7:"jul", 8:"ago", 9:"sep", 10:"oct", 11:"nov", 12:"dic"}


# ── Lógica de datos ───────────────────────────────────────────────────────────
def leer_venta_chips(ruta: str) -> pd.DataFrame:
    """Lee el archivo Venta Chips y devuelve un DataFrame limpio."""
    xls = pd.read_excel(ruta, sheet_name=None, header=None)
    
    # Buscar la hoja con datos reales (BASE o Hoja1 con encabezado en fila 1/2)
    for nombre, df in xls.items():
        # Si tiene columna 'vendedor' directo → BASE
        cols = [str(c).strip().lower() for c in df.iloc[0]]
        if "vendedor" in cols:
            df.columns = df.iloc[0]
            df = df[1:].reset_index(drop=True)
            df.columns = [str(c).strip().lower() for c in df.columns]
            if "vendedor" in df.columns and "fecha" in df.columns:
                return _limpiar_df(df)
        # Buscar fila encabezado
        for i, row in df.iterrows():
            vals = [str(v).strip().lower() for v in row]
            if "vendedor" in vals and "fecha" in vals:
                df.columns = df.iloc[i]
                df = df[i+1:].reset_index(drop=True)
                df.columns = [str(c).strip().lower() for c in df.columns]
                return _limpiar_df(df)
    raise ValueError("No se encontró la hoja con columnas 'fecha' y 'vendedor'.")


def _limpiar_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.dropna(subset=["vendedor", "fecha"]).copy()
    df["fecha"]    = pd.to_datetime(df["fecha"], errors="coerce")
    df["vendedor"] = df["vendedor"].astype(str).str.strip()
    df["cliente"]  = df["cliente"].astype(str).str.strip() if "cliente" in df.columns else "Sin cliente"
    df = df.dropna(subset=["fecha"])
    df["mes"]  = df["fecha"].dt.month
    df["anio"] = df["fecha"].dt.year
    df["mes_lbl"] = df["mes"].map(MESES_ES)
    return df


def obtener_vendedores(df: pd.DataFrame) -> list:
    return sorted(df["vendedor"].unique().tolist())


def datos_vendedor(df: pd.DataFrame, vendedor: str) -> pd.DataFrame:
    return df[df["vendedor"] == vendedor].copy()


def resumen_mensual(df_v: pd.DataFrame) -> pd.DataFrame:
    """Pivot: filas=cliente, columnas=mes_lbl, valores=count."""
    meses_orden = [MESES_ES[m] for m in sorted(df_v["mes"].unique())]
    pivot = (df_v.groupby(["cliente", "mes_lbl"])
               .size()
               .unstack(fill_value=0)
               .reindex(columns=meses_orden, fill_value=0))
    pivot["TOTAL"]    = pivot.sum(axis=1)
    pivot["PROMEDIO"] = pivot[meses_orden].mean(axis=1).round(2)
    return pivot.sort_values("TOTAL", ascending=False)


def generar_resumen_xlsx(df: pd.DataFrame, ruta_salida: str):
    """Genera RESUMEN.xlsx con la estructura exacta del archivo original."""
    meses_unicos = sorted(df["mes"].unique())
    meses_lbls   = [MESES_ES[m] for m in meses_unicos]
    vendedores   = obtener_vendedores(df)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RESUMEN"

    # ── Estilos ──
    hdr_font   = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill   = PatternFill("solid", fgColor="1F4E79")
    vend_fill  = PatternFill("solid", fgColor="2E75B6")
    vend_font  = Font(bold=True, color="FFFFFF", size=10)
    alt_fill   = PatternFill("solid", fgColor="D6E4F0")
    center     = Alignment(horizontal="center", vertical="center")
    left_al    = Alignment(horizontal="left",   vertical="center")
    thin       = Side(style="thin", color="AAAAAA")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Cabecera ──
    headers = ["VENDEDOR", "Etiquetas de fila", "INVENTARIO"] + meses_lbls + ["PROMEDIO"]
    for col_i, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_i, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        cell.border    = border

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 12
    for i in range(len(meses_lbls)):
        col_letter = openpyxl.utils.get_column_letter(4 + i)
        ws.column_dimensions[col_letter].width = 8
    ws.column_dimensions[openpyxl.utils.get_column_letter(4+len(meses_lbls))].width = 12

    fila = 2
    for vend in vendedores:
        df_v   = datos_vendedor(df, vend)
        pivot  = resumen_mensual(df_v)

        # Fila agrupadora del vendedor
        ws.cell(row=fila, column=1, value=vend).font = vend_font
        ws.cell(row=fila, column=1).fill      = vend_fill
        ws.cell(row=fila, column=1).alignment = left_al
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=len(headers))
        for c in range(1, len(headers)+1):
            ws.cell(row=fila, column=c).border = border
        fila += 1

        for idx_row, (cliente, row_data) in enumerate(pivot.iterrows()):
            fill = alt_fill if idx_row % 2 == 0 else None
            ws.cell(row=fila, column=1, value=vend).alignment = left_al
            ws.cell(row=fila, column=2, value=cliente).alignment = left_al
            ws.cell(row=fila, column=3, value=0).alignment = center  # INVENTARIO (sin fuente en Venta Chips)
            for m_i, mes in enumerate(meses_lbls):
                val = int(row_data.get(mes, 0))
                ws.cell(row=fila, column=4+m_i, value=val).alignment = center
            ws.cell(row=fila, column=4+len(meses_lbls), value=round(float(row_data["PROMEDIO"]), 6)).alignment = center
            for c in range(1, len(headers)+1):
                cell = ws.cell(row=fila, column=c)
                cell.border = border
                if fill:
                    cell.fill = fill
            fila += 1

    # ── NUEVO CÓDIGO: Fila de Total General ──
    # 1. Título y formato
    ws.cell(row=fila, column=1, value="Total general").font = hdr_font
    ws.cell(row=fila, column=1).fill = hdr_fill
    ws.cell(row=fila, column=1).alignment = left_al
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=2)

    # 2. Inventario Total (Se mantiene en 0)
    ws.cell(row=fila, column=3, value=0).font = hdr_font
    ws.cell(row=fila, column=3).fill = hdr_fill
    ws.cell(row=fila, column=3).alignment = center

    # 3. Sumatorias de cada mes con fórmulas de Excel (D2:D...)
    for m_i in range(len(meses_lbls)):
        col_letter = openpyxl.utils.get_column_letter(4 + m_i)
        formula = f"=SUM({col_letter}2:{col_letter}{fila-1})"
        
        ws.cell(row=fila, column=4+m_i, value=formula).font = hdr_font
        ws.cell(row=fila, column=4+m_i).fill = hdr_fill
        ws.cell(row=fila, column=4+m_i).alignment = center

    # 4. Sumatoria de la columna PROMEDIO
    col_prom = openpyxl.utils.get_column_letter(4 + len(meses_lbls))
    formula_prom = f"=SUM({col_prom}2:{col_prom}{fila-1})"
    
    ws.cell(row=fila, column=4+len(meses_lbls), value=formula_prom).font = hdr_font
    ws.cell(row=fila, column=4+len(meses_lbls)).fill = hdr_fill
    ws.cell(row=fila, column=4+len(meses_lbls)).alignment = center

    # 5. Aplicar bordes a toda la fila final
    for c in range(1, len(headers)+1):
        ws.cell(row=fila, column=c).border = border
    # ─────────────────────────────────────────

    # Hoja BASE
    ws_base = wb.create_sheet("BASE")
    cols_base = ["fecha","carrier","vendedor","cliente","monto","iccid","dn","producto","plaza"]
    cols_exist = [c for c in cols_base if c in df.columns]
    for ci, c in enumerate(cols_exist, 1):
        cell = ws_base.cell(row=1, column=ci, value=c.upper())
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        cell.border    = border
    for ri, (_, row) in enumerate(df[cols_exist].iterrows(), 2):
        for ci, c in enumerate(cols_exist, 1):
            ws_base.cell(row=ri, column=ci, value=row[c]).border = border

    wb.save(ruta_salida)


# ── Interfaz principal ────────────────────────────────────────────────────────
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Dashboard Ventas Chips ATT")
        self.geometry("1400x820")
        self.minsize(1100, 650)
        self.configure(fg_color=BG_DARK)

        self.df_full: pd.DataFrame | None = None
        self.vendedor_actual: str | None  = None
        self.canvas_fig = None
        self._construir_ui()

    # ── Layout ────────────────────────────────────────────────────────────────
    def _construir_ui(self):
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # ── Panel izquierdo ──
        self.panel_izq = ctk.CTkFrame(self, width=270, corner_radius=0, fg_color=BG_MID)
        self.panel_izq.grid(row=0, column=0, sticky="nsew")
        self.panel_izq.grid_propagate(False)
        self.panel_izq.grid_rowconfigure(4, weight=1)

        lbl_title = ctk.CTkLabel(self.panel_izq, text="📊 Dashboard Chips",
                                  font=("Segoe UI", 16, "bold"), text_color=TEXT_CLR)
        lbl_title.grid(row=0, column=0, padx=16, pady=(18, 4), sticky="w")

        self.btn_cargar = ctk.CTkButton(self.panel_izq, text="📂  Cargar archivo",
                                         command=self._cargar_archivo,
                                         fg_color=ACCENT, hover_color="#145082",
                                         font=("Segoe UI", 12, "bold"), height=36)
        self.btn_cargar.grid(row=1, column=0, padx=14, pady=(6, 2), sticky="ew")

        self.lbl_archivo = ctk.CTkLabel(self.panel_izq, text="Sin archivo cargado",
                                         font=("Segoe UI", 9), text_color="#888888",
                                         wraplength=240)
        self.lbl_archivo.grid(row=2, column=0, padx=14, pady=(0, 8), sticky="w")

        self.entry_buscar = ctk.CTkEntry(self.panel_izq, placeholder_text="🔍  Buscar vendedor...",
                                          font=("Segoe UI", 11), height=34)
        self.entry_buscar.grid(row=3, column=0, padx=14, pady=(0, 6), sticky="ew")
        self.entry_buscar.bind("<KeyRelease>", self._filtrar_vendedores)

        self.scroll_vendedores = ctk.CTkScrollableFrame(self.panel_izq, fg_color=BG_MID,
                                                         label_text="", corner_radius=8)
        self.scroll_vendedores.grid(row=4, column=0, padx=10, pady=(0, 10), sticky="nsew")

        self.btn_resumen = ctk.CTkButton(self.panel_izq, text="📄  Generar RESUMEN",
                                          command=self._generar_resumen,
                                          fg_color="#2E7D32", hover_color="#1B5E20",
                                          font=("Segoe UI", 12, "bold"), height=36,
                                          state="disabled")
        self.btn_resumen.grid(row=5, column=0, padx=14, pady=(6, 14), sticky="ew")

        # ── Panel derecho ──
        self.panel_der = ctk.CTkFrame(self, corner_radius=0, fg_color=BG_DARK)
        self.panel_der.grid(row=0, column=1, sticky="nsew")
        self.panel_der.grid_rowconfigure(1, weight=1)
        self.panel_der.grid_columnconfigure(0, weight=1)

        self._mostrar_bienvenida()

    def _mostrar_bienvenida(self):
        for w in self.panel_der.winfo_children():
            w.destroy()
        frm = ctk.CTkFrame(self.panel_der, fg_color="transparent")
        frm.place(relx=0.5, rely=0.5, anchor="center")
        ctk.CTkLabel(frm, text="📱", font=("Segoe UI", 64)).pack()
        ctk.CTkLabel(frm, text="Dashboard de Ventas de Chips ATT",
                      font=("Segoe UI", 22, "bold"), text_color=TEXT_CLR).pack(pady=(8, 4))
        ctk.CTkLabel(frm, text="Carga un archivo Venta_Chips para comenzar",
                      font=("Segoe UI", 13), text_color="#888888").pack()

    # ── Carga de archivo ──────────────────────────────────────────────────────
    def _cargar_archivo(self):
        ruta = filedialog.askopenfilename(
            title="Seleccionar archivo Venta_Chips",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")]
        )
        if not ruta:
            return
        self.btn_cargar.configure(text="⏳ Cargando...", state="disabled")

        result = {"df": None, "error": None}

        def _worker():
            try:
                result["df"] = leer_venta_chips(ruta)
            except Exception as e:
                result["error"] = str(e)
            self.after(0, lambda: _done())

        def _done():
            self.btn_cargar.configure(text="📂  Cargar archivo", state="normal")
            if result["error"]:
                self.lbl_archivo.configure(text="❌ Error al cargar", text_color="#f44336")
                messagebox.showerror("Error", f"No se pudo leer el archivo:\n\n{result['error']}")
            else:
                self.df_full = result["df"]
                nombre = os.path.basename(ruta)
                self.lbl_archivo.configure(text=f"✅ {nombre}", text_color="#4CAF50")
                self.btn_resumen.configure(state="normal")
                self._poblar_vendedores(obtener_vendedores(self.df_full))
                self._mostrar_bienvenida()

        threading.Thread(target=_worker, daemon=True).start()

    # ── Lista de vendedores ───────────────────────────────────────────────────
    def _poblar_vendedores(self, vendedores: list):
        for w in self.scroll_vendedores.winfo_children():
            w.destroy()
        self._botones_vendedor = {}
        for v in vendedores:
            btn = ctk.CTkButton(
                self.scroll_vendedores, text=v,
                font=("Segoe UI", 10), anchor="w", height=32,
                fg_color="transparent", hover_color=BG_CARD,
                text_color=TEXT_CLR, border_width=0,
                command=lambda vend=v: self._seleccionar_vendedor(vend)
            )
            btn.pack(fill="x", padx=4, pady=2)
            self._botones_vendedor[v] = btn

    def _filtrar_vendedores(self, event=None):
        if not self.df_full is not None and hasattr(self, "_botones_vendedor"):
            return
        if not hasattr(self, "_botones_vendedor"):
            return
        texto = self.entry_buscar.get().lower()
        for nombre, btn in self._botones_vendedor.items():
            if texto in nombre.lower():
                btn.pack(fill="x", padx=4, pady=2)
            else:
                btn.pack_forget()

    def _seleccionar_vendedor(self, vendedor: str):
        # Resaltar botón seleccionado
        if hasattr(self, "_botones_vendedor"):
            for n, b in self._botones_vendedor.items():
                b.configure(fg_color=BG_CARD if n == vendedor else "transparent")
        self.vendedor_actual = vendedor
        self._mostrar_graficas(vendedor)

    # ── Gráficas ──────────────────────────────────────────────────────────────
    def _mostrar_graficas(self, vendedor: str):
        for w in self.panel_der.winfo_children():
            w.destroy()

        df_v  = datos_vendedor(self.df_full, vendedor)
        pivot = resumen_mensual(df_v)
        meses = [c for c in pivot.columns if c not in ("TOTAL", "PROMEDIO")]

        # ── Encabezado ──
        frm_hdr = ctk.CTkFrame(self.panel_der, fg_color=BG_MID, corner_radius=0, height=56)
        frm_hdr.grid(row=0, column=0, sticky="ew")
        frm_hdr.grid_propagate(False)
        ctk.CTkLabel(frm_hdr, text=f"🧑‍💼  {vendedor}",
                      font=("Segoe UI", 14, "bold"), text_color=TEXT_CLR).pack(side="left", padx=16, pady=12)

        total_chips = int(pivot["TOTAL"].sum())
        prom_global = round(float(pivot["PROMEDIO"].mean()), 2)
        ctk.CTkLabel(frm_hdr, text=f"Total chips: {total_chips:,}   |   Promedio global: {prom_global}",
                      font=("Segoe UI", 11), text_color="#aaaaaa").pack(side="right", padx=20)

        # ── Controles: selector de vista ──
        frm_ctrl = ctk.CTkFrame(self.panel_der, fg_color=BG_DARK, height=44)
        frm_ctrl.grid(row=1, column=0, sticky="ew", padx=16, pady=(8, 0))
        self._vista = ctk.StringVar(value="barras")
        for txt, val in [("Barras por cliente", "barras"), ("Por mes (total)", "meses"), ("Top 15 clientes", "top15")]:
            ctk.CTkRadioButton(frm_ctrl, text=txt, variable=self._vista, value=val,
                                font=("Segoe UI", 10), command=lambda: self._refrescar_grafica(pivot, meses)
                                ).pack(side="left", padx=12, pady=8)

        # ── Área de figura ──
        self._frm_fig = ctk.CTkFrame(self.panel_der, fg_color=BG_DARK)
        self._frm_fig.grid(row=2, column=0, sticky="nsew", padx=0, pady=0)
        self.panel_der.grid_rowconfigure(2, weight=1)

        self._pivot_actual = pivot
        self._meses_actual = meses
        self._refrescar_grafica(pivot, meses)

    def _refrescar_grafica(self, pivot, meses):
        for w in self._frm_fig.winfo_children():
            w.destroy()
        vista = self._vista.get()
        if vista == "barras":
            self._fig_barras(pivot, meses)
        elif vista == "meses":
            self._fig_meses(pivot, meses)
        else:
            self._fig_top15(pivot, meses)

    def _estilo_fig(self):
        fig, ax = plt.subplots(figsize=(10, 6))
        fig.patch.set_facecolor("#1a1a2e")
        ax.set_facecolor("#0f1729")
        ax.tick_params(colors=TEXT_CLR, labelsize=8)
        ax.xaxis.label.set_color(TEXT_CLR)
        ax.yaxis.label.set_color(TEXT_CLR)
        ax.title.set_color(TEXT_CLR)
        for spine in ax.spines.values():
            spine.set_edgecolor("#333355")
        ax.yaxis.set_major_locator(mticker.MaxNLocator(integer=True))
        ax.grid(axis="y", color="#333355", linestyle="--", linewidth=0.5)
        return fig, ax

    def _fig_barras(self, pivot, meses):
        # Mostrar barras agrupadas por mes para cada cliente (top 20)
        top = pivot.head(20)
        n_clientes = len(top)
        n_meses    = len(meses)
        import numpy as np

        fig, ax = self._estilo_fig()
        x = np.arange(n_clientes)
        ancho = 0.8 / max(n_meses, 1)

        for i, mes in enumerate(meses):
            vals = top[mes].values
            offset = (i - n_meses/2 + 0.5) * ancho
            bars = ax.bar(x + offset, vals, width=ancho, label=mes,
                          color=MONTH_COLORS[i % len(MONTH_COLORS)], alpha=0.88)

        # Línea de promedio
        ax.plot(x, top["PROMEDIO"].values, color="#FFD700", linewidth=2,
                marker="D", markersize=5, label="Promedio", zorder=5)

        ax.set_xticks(x)
        ax.set_xticklabels(top.index.tolist(), rotation=35, ha="right", fontsize=7.5)
        ax.set_title("Chips por cliente y mes", fontsize=12, pad=10)
        ax.set_ylabel("Chips")
        ax.legend(loc="upper right", fontsize=8, facecolor="#1a1a2e",
                  labelcolor=TEXT_CLR, edgecolor="#444466")
        fig.tight_layout()
        self._embeber_fig(fig)

    def _fig_meses(self, pivot, meses):
        totales = [pivot[m].sum() for m in meses]
        prom    = sum(totales) / len(totales) if totales else 0
        import numpy as np

        fig, ax = self._estilo_fig()
        bars = ax.bar(meses, totales,
                      color=[MONTH_COLORS[i % len(MONTH_COLORS)] for i in range(len(meses))],
                      alpha=0.88, width=0.55)
        ax.axhline(prom, color="#FFD700", linestyle="--", linewidth=1.8, label=f"Promedio: {prom:.1f}")
        for bar, val in zip(bars, totales):
            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.5,
                    str(int(val)), ha="center", va="bottom", fontsize=9, color=TEXT_CLR)
        ax.set_title("Total chips por mes", fontsize=12, pad=10)
        ax.set_ylabel("Chips")
        ax.legend(fontsize=9, facecolor="#1a1a2e", labelcolor=TEXT_CLR, edgecolor="#444466")
        fig.tight_layout()
        self._embeber_fig(fig)

    def _fig_top15(self, pivot, meses):
        top = pivot.head(15).copy()
        import numpy as np

        fig, ax = self._estilo_fig()
        y = np.arange(len(top))
        ax.barh(y, top["TOTAL"].values, color=ACCENT, alpha=0.85, height=0.55)
        ax.plot(top["PROMEDIO"].values * (len(meses) / max(len(meses), 1)),
                y, "o", color="#FFD700", zorder=5, label="Promedio×meses")
        ax.set_yticks(y)
        ax.set_yticklabels(top.index.tolist(), fontsize=8)
        for i, (_, row) in enumerate(top.iterrows()):
            ax.text(row["TOTAL"] + 0.3, i, str(int(row["TOTAL"])),
                    va="center", fontsize=8, color=TEXT_CLR)
        ax.set_title("Top 15 clientes (total chips)", fontsize=12, pad=10)
        ax.set_xlabel("Chips totales")
        ax.invert_yaxis()
        fig.tight_layout()
        self._embeber_fig(fig)

    def _embeber_fig(self, fig):
        canvas = FigureCanvasTkAgg(fig, master=self._frm_fig)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)
        plt.close(fig)

    # ── Generación de RESUMEN ─────────────────────────────────────────────────
    def _generar_resumen(self):
        if self.df_full is None:
            messagebox.showwarning("Sin datos", "Primero carga un archivo Venta_Chips.")
            return
        ruta = filedialog.asksaveasfilename(
            title="Guardar RESUMEN como...",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="RESUMEN_Generado.xlsx"
        )
        if not ruta:
            return

        # Capturar df en local para pasar al hilo sin referencias a tkinter
        df_snapshot = self.df_full.copy()
        self.btn_resumen.configure(text="⏳ Generando...", state="disabled")

        result = {"error": None}

        # ── NUEVO: Forzar recolección de basura en el hilo principal ──
        # Esto limpia rastros del filedialog y gráficas viejas antes
        # de irnos al hilo secundario, evitando el RuntimeError.
        gc.collect()
        # ──────────────────────────────────────────────────────────────

        def _worker():
            try:
                generar_resumen_xlsx(df_snapshot, ruta)
            except Exception as e:
                result["error"] = str(e)
            # Notificar al hilo principal via after
            self.after(0, lambda: _done())

        def _done():
            self.btn_resumen.configure(text="📄  Generar RESUMEN", state="normal")
            if result["error"]:
                messagebox.showerror("Error al generar",
                                     f"No se pudo generar el RESUMEN:\n\n{result['error']}")
            else:
                messagebox.showinfo("¡Listo!",
                                    f"RESUMEN generado exitosamente:\n{ruta}")

        threading.Thread(target=_worker, daemon=True).start()


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = App()
    app.mainloop()